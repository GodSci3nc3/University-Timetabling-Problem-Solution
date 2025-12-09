"""
Generador de Excel con 3 hojas siguiendo el formato de referencia:
1. Matriz ITI: Asignación de profesores a materias
2. ITI: Horarios visuales por grupo  
3. Disponibilidad: Disponibilidad de profesores por slot
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Any
from collections import defaultdict


class GeneradorExcelResultado:
    """Genera archivo Excel con los resultados del horario en 3 hojas"""
    
    def __init__(self, datos_entrada: Dict, horario_generado: Dict):
        """
        Args:
            datos_entrada: Diccionario con grupos, materias, profesores
            horario_generado: Horario del backend (estructura: {grupo: {dia: {slot: {materia, profesor}}}})
        """
        self.datos = datos_entrada
        self.horario = horario_generado
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Eliminar hoja por defecto
        
        # Estilos comunes
        self.font_header = Font(name='Arial', size=10, bold=True)
        self.font_normal = Font(name='Arial', size=9)
        self.font_small = Font(name='Arial', size=8)
        self.align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Colores
        self.fill_header = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.fill_gray = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        self.fill_light = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        # Bordes
        thin_border = Side(style='thin', color='000000')
        self.border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    def generar(self, ruta_salida: str):
        """Genera las 3 hojas y guarda el archivo"""
        self._generar_hoja_matriz()
        self._generar_hoja_horarios()
        self._generar_hoja_disponibilidad()
        self.wb.save(ruta_salida)
        return ruta_salida
    
    def _generar_hoja_matriz(self):
        """Hoja 1: Matriz de asignación profesores-materias"""
        ws = self.wb.create_sheet("Matriz ITI")
        
        # Obtener lista de profesores únicos
        profesores = sorted(set(self.datos['profesores'].keys()))
        
        # Headers
        ws['A2'] = 'grupos'
        ws['B2'] = 'Horas x materia'
        ws['C2'] = 'Horas x Semana'
        ws['D2'] = 'Resta'
        
        # Nombres de profesores en columnas
        col_offset = 5  # Empieza en columna F (índice 6)
        for idx, profesor in enumerate(profesores):
            col = get_column_letter(col_offset + idx + 1)
            ws[f'{col}2'] = profesor
            ws[f'{col}2'].font = self.font_header
            ws[f'{col}2'].alignment = self.align_center
            ws[f'{col}2'].fill = self.fill_header
        
        # Aplicar estilos a headers
        for cell in ['A2', 'B2', 'C2', 'D2']:
            ws[cell].font = self.font_header
            ws[cell].alignment = self.align_center
            ws[cell].fill = self.fill_header
        
        # Fila de totales (fila 3)
        ws['A3'] = 'Horas Asignadas'
        ws['A3'].font = self.font_header
        for idx in range(len(profesores)):
            col = get_column_letter(col_offset + idx + 1)
            # Fórmula suma (se actualizará al final según filas)
            ws[f'{col}3'] = f'=SUM({col}4:{col}100)'
        
        # Procesar asignaciones por grupo y materia
        # Estructura del horario: {grupo: {dia: {slot: {materia, profesor}}}}
        
        # Agrupar por grupo-materia
        matriz_datos = defaultdict(lambda: defaultdict(int))  # [grupo-materia][profesor] = horas
        horas_por_grupo_materia = defaultdict(int)
        
        for grupo, dias_data in self.horario.items():
            for dia, slots_data in dias_data.items():
                for slot, asig_data in slots_data.items():
                    if asig_data:
                        materia = asig_data.get('materia', '')
                        profesor = asig_data.get('profesor', '')
                        
                        grupo_materia = f"{grupo} - {materia}"
                        matriz_datos[grupo_materia][profesor] += 1
                        horas_por_grupo_materia[grupo_materia] += 1
        
        # Escribir datos por fila
        fila_actual = 4
        grupos_vistos = set()
        
        for grupo_materia in sorted(matriz_datos.keys()):
            grupo = grupo_materia.split(' - ')[0]
            
            # Header de cuatrimestre/grupo (si es nuevo)
            if grupo not in grupos_vistos:
                ws[f'A{fila_actual}'] = grupo
                ws[f'A{fila_actual}'].font = Font(bold=True, size=10)
                ws[f'A{fila_actual}'].fill = self.fill_gray
                fila_actual += 1
                grupos_vistos.add(grupo)
            
            # Datos de grupo-materia
            ws[f'A{fila_actual}'] = grupo_materia
            ws[f'B{fila_actual}'] = horas_por_grupo_materia[grupo_materia]
            
            # Horas por profesor
            for idx, profesor in enumerate(profesores):
                col = get_column_letter(col_offset + idx + 1)
                horas = matriz_datos[grupo_materia].get(profesor, 0)
                if horas > 0:
                    ws[f'{col}{fila_actual}'] = horas
                    ws[f'{col}{fila_actual}'].alignment = self.align_center
            
            fila_actual += 1
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        for idx in range(len(profesores)):
            col = get_column_letter(col_offset + idx + 1)
            ws.column_dimensions[col].width = 12
    
    def _generar_hoja_horarios(self):
        """Hoja 2: Horarios visuales por grupo con días L-M-Mi-J-V"""
        ws = self.wb.create_sheet("ITI")
        
        # Obtener grupos únicos
        grupos = sorted(set(self.datos['grupos'].keys()))
        
        # Slots horarios (extraer del resultado)
        slots_horarios = self._obtener_slots_unicos()
        dias = ['L', 'M', 'Mi', 'J', 'V']
        
        # Generar columnas por grupo
        col_actual = 1
        for grupo in grupos:
            # Header del grupo
            col_letra = get_column_letter(col_actual)
            ws[f'{col_letra}2'] = grupo
            ws[f'{col_letra}2'].font = self.font_header
            ws[f'{col_letra}2'].fill = self.fill_header
            ws[f'{col_letra}2'].alignment = self.align_center
            ws.merge_cells(f'{col_letra}2:{get_column_letter(col_actual + 4)}2')
            
            # Días de la semana
            for idx_dia, dia in enumerate(dias):
                col_dia = get_column_letter(col_actual + idx_dia)
                ws[f'{col_dia}3'] = dia
                ws[f'{col_dia}3'].font = self.font_header
                ws[f'{col_dia}3'].alignment = self.align_center
                ws[f'{col_dia}3'].fill = self.fill_light
            
            # Llenar horarios
            fila_actual = 4
            for slot in slots_horarios:
                # Primera columna: hora
                if col_actual == 1:
                    ws[f'A{fila_actual}'] = slot
                    ws[f'A{fila_actual}'].font = self.font_small
                    ws[f'A{fila_actual}'].alignment = self.align_center
                
                # Buscar asignaciones para este grupo-slot
                for idx_dia, dia in enumerate(dias):
                    col_dia = get_column_letter(col_actual + idx_dia)
                    asig_texto = self._buscar_asignacion_grupo_dia_slot(grupo, dia, slot)
                    ws[f'{col_dia}{fila_actual}'] = asig_texto
                    ws[f'{col_dia}{fila_actual}'].font = self.font_small
                    ws[f'{col_dia}{fila_actual}'].alignment = self.align_center
                    ws[f'{col_dia}{fila_actual}'].border = self.border
                
                fila_actual += 1
            
            col_actual += 6  # Espacio entre grupos (5 días + 1 separación)
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        for grupo in grupos:
            for i in range(5):
                col = get_column_letter(col_actual - len(grupos) * 6 + i)
                ws.column_dimensions[col].width = 15
    
    def _generar_hoja_disponibilidad(self):
        """Hoja 3: Disponibilidad de profesores por slot"""
        ws = self.wb.create_sheet("Disponibilidad")
        
        # Obtener profesores y slots
        profesores = sorted(set(self.datos['profesores'].keys()))
        slots_horarios = self._obtener_slots_unicos()
        dias = ['L', 'M', 'Mi', 'J', 'V']
        
        # Headers de profesores
        col_actual = 2
        for profesor in profesores:
            col_letra = get_column_letter(col_actual)
            # Nombre del profesor (fila 1)
            ws[f'{col_letra}1'] = profesor
            ws[f'{col_letra}1'].font = self.font_header
            ws[f'{col_letra}1'].alignment = self.align_center
            ws.merge_cells(f'{col_letra}1:{get_column_letter(col_actual + 4)}1')
            
            # Días (fila 2)
            for idx_dia, dia in enumerate(dias):
                col_dia = get_column_letter(col_actual + idx_dia)
                ws[f'{col_dia}2'] = dia
                ws[f'{col_dia}2'].font = self.font_small
                ws[f'{col_dia}2'].alignment = self.align_center
                ws[f'{col_dia}2'].fill = self.fill_light
            
            col_actual += 6  # 5 días + 1 separación
        
        # Llenar disponibilidad
        fila_actual = 3
        for slot in slots_horarios:
            # Primera columna: hora
            ws[f'A{fila_actual}'] = slot
            ws[f'A{fila_actual}'].font = self.font_small
            ws[f'A{fila_actual}'].alignment = self.align_center
            
            # Por cada profesor
            col_actual = 2
            for profesor in profesores:
                for idx_dia, dia in enumerate(dias):
                    col_dia = get_column_letter(col_actual + idx_dia)
                    grupo_asignado = self._buscar_grupo_profesor_dia_slot(profesor, dia, slot)
                    ws[f'{col_dia}{fila_actual}'] = grupo_asignado
                    ws[f'{col_dia}{fila_actual}'].font = self.font_small
                    ws[f'{col_dia}{fila_actual}'].alignment = self.align_center
                    ws[f'{col_dia}{fila_actual}'].border = self.border
                
                col_actual += 6
            
            fila_actual += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 12
        for i in range(2, col_actual):
            ws.column_dimensions[get_column_letter(i)].width = 10
    
    def _obtener_slots_unicos(self) -> List[str]:
        """Extrae lista única de slots horarios del resultado"""
        slots = set()
        # Estructura: {grupo: {dia: {slot: {materia, profesor}}}}
        for grupo, dias_data in self.horario.items():
            for dia, slots_data in dias_data.items():
                for slot in slots_data.keys():
                    slots.add(slot)
        
        # Ordenar por hora de inicio
        def slot_key(s):
            hora = s.split('-')[0]
            h, m = hora.split(':')
            return int(h) * 60 + int(m)
        
        return sorted(list(slots), key=slot_key)
    
    def _buscar_asignacion_grupo_dia_slot(self, grupo: str, dia: str, slot: str) -> str:
        """Busca la asignación para un grupo-día-slot específico"""
        # Mapeo de días
        dia_map = {'L': 'Lunes', 'M': 'Martes', 'Mi': 'Miercoles', 'J': 'Jueves', 'V': 'Viernes'}
        dia_completo = dia_map.get(dia, dia)
        
        # Estructura: {grupo: {dia: {slot: {materia, profesor}}}}
        if grupo in self.horario:
            # Buscar con nombre completo o sin acento
            for dia_key in [dia_completo, dia_completo.replace('é', 'e')]:
                if dia_key in self.horario[grupo]:
                    if slot in self.horario[grupo][dia_key]:
                        asig_data = self.horario[grupo][dia_key][slot]
                        if asig_data:
                            materia = asig_data.get('materia', '')
                            profesor = asig_data.get('profesor', '')
                            return f"{materia}\n{profesor}"
        
        return ""
    
    def _buscar_grupo_profesor_dia_slot(self, profesor: str, dia: str, slot: str) -> str:
        """Busca el grupo asignado a un profesor en un día-slot específico"""
        dia_map = {'L': 'Lunes', 'M': 'Martes', 'Mi': 'Miercoles', 'J': 'Jueves', 'V': 'Viernes'}
        dia_completo = dia_map.get(dia, dia)
        
        # Estructura: {grupo: {dia: {slot: {materia, profesor}}}}
        for grupo, dias_data in self.horario.items():
            # Buscar con nombre completo o sin acento
            for dia_key in [dia_completo, dia_completo.replace('é', 'e')]:
                if dia_key in dias_data:
                    if slot in dias_data[dia_key]:
                        asig_data = dias_data[dia_key][slot]
                        if asig_data and asig_data.get('profesor') == profesor:
                            return grupo
        
        return ""
